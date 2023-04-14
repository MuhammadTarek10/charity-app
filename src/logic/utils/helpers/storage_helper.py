import openpyxl as op

from types import FunctionType
from src.logic.config.strings import Strings
from src.logic.data.database import Database


from src.logic.utils.helpers.pops import Pops
from src.logic.data.models import *


class StorageHelper:
    instance = None

    @staticmethod
    def getInstance(
        adjustTable: FunctionType = None,
    ) -> "StorageHelper":
        if StorageHelper.instance is None:
            StorageHelper.instance = StorageHelper(
                adjustTable=adjustTable,
            )
        return StorageHelper.instance

    def setPops(self, pops: Pops) -> None:
        self.pops = pops

    def __init__(
        self,
        pops: Pops = None,
        adjustTable: FunctionType = None,
    ) -> None:
        self.pops = pops
        self.database: Database = Database(db_url=Config.DATABASE_URL)
        self.database.insertImportantCaseTypes()
        self.adjustTable: FunctionType = adjustTable
        self.data: list[tuple] = None
        self.sheet: op.Worksheet = None

    def getDataExcel(self) -> None:
        filepath = "test.xlsx"  # TODO: Remove
        try:
            workbook: op.Workbook = op.load_workbook(filepath)
            self.sheet = workbook.active
            self.data = list(self.sheet.values)
            if len(self.data) > 1:
                self.adjustTable(self.data, self.sheet.max_row, self.sheet.max_column)
        except:
            pass

    def totalMoney(self) -> float:
        return self.database.totalMoney()

    # *Cases
    def getAllCases(self) -> list:
        return self.database.getAllCases()

    def getCaseByName(self, name: str) -> Case:
        return self.database.getCaseByName(name)

    def insertCase(self, data: dict) -> None:
        self.database.insertCase(data)

    # *Cases Types
    def getAllCaseTypes(self) -> list:
        return self.database.getCaseAllTypes()

    def insertCaseType(self, data: dict) -> bool:
        if self.database.getCaseTypeByName(data[Config.CASES_TYPES_NAME]) is None:
            self.database.insertCaseType(data)
            return True
        return False

    def deleteCaseType(self, name: str) -> bool:
        if self.database.getCaseTypeByName(name) is not None:
            self.database.deleteCaseType(name)
            return True
        return False

    # *Invoices
    def getAllInvoices(self) -> list:
        return self.database.getAllInvoices()

    def getInvoiceByName(self, name: str) -> Invoice:
        return self.database.getInvoiceByName(name)

    def insertInvoice(self, data: dict) -> None:
        invoice = {
            Config.INVOICES_DATE: data[Config.INVOICES_DATE],
            Config.INVOICES_VALUE: data[Config.INVOICES_VALUE],
            Config.INVOICES_CASE_ID: data[Config.INVOICES_CASE_ID],
        }
        try:
            item = {
                Config.ITEMS_NAME: data[Config.ITEMS_NAME],
                Config.ITEMS_UNIT: data[Config.ITEMS_UNIT],
                Config.ITEMS_QUANTITY: int(data[Config.ITEMS_QUANTITY]) * -1,
                Config.ITEM_IS_IN_STORAGE: False,
            }
            # get id of item and make it in invoice
            invoice[Config.INVOICES_ITEM_ID] = self.database.insertItem(item)
        except Exception as e:
            print(e)
        self.database.insertInvoice(invoice)

    def insertDonator(self, data: dict) -> None:
        self.database.insertDonator(data)

    # *Donations
    def getAllDonations(self) -> list:
        return self.database.getAllDonations()

    def insertDonation(self, data: dict) -> None:
        donation = {
            Config.DONATIONS_DONATOR_NAME: data[Config.DONATIONS_DONATOR_NAME],
            Config.DONATION_DATE: data[Config.DONATION_DATE],
            Config.DONATIONS_VALUE: data[Config.DONATIONS_VALUE],
            Config.DONATIONS_ITEM_TYPE: data[Config.DONATIONS_ITEM_TYPE],
        }
        try:
            item = {
                Config.ITEMS_NAME: data[Config.ITEMS_NAME],
                Config.ITEMS_UNIT: data[Config.ITEMS_UNIT],
                Config.ITEMS_QUANTITY: data[Config.ITEMS_QUANTITY],
                Config.ITEMS_PRICE: data[Config.ITEMS_PRICE],
                Config.ITEM_IS_IN_STORAGE: True,
            }
            # get id of item and make it in donation
            donation[Config.DONATIONS_ITEM_ID] = self.database.insertItem(item)
        except:
            pass
        self.database.insertDonation(donation)

    # *Items
    def getAllItems(self) -> list:
        return self.database.getAllItems()

    def getItemByName(self, name: str) -> Item:
        return self.database.getItemByName(name)

    def insertItem(self, data: dict) -> None:
        self.database.insertItem(data)
